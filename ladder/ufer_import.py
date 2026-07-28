"""Import the existing ranking spreadsheet as a starting point.

The workbook carries forward final standings, not the full event chain: the
`Elo Algorithm` sheets only hold the most recent calculation, and older
games live in per-tournament bracket layouts that differ from sheet to
sheet. Recomputing from season one is not possible without manual work per
tournament.

So current ratings are taken as starting values and everything from then on
is calculated by `ratings.recompute`. Future numbers are verifiable without
having to rewrite the past.

What the sheet does not have is Steam IDs — only display names, which need
not match the Steam persona. That link comes from the recorder, because the
game log holds name and SteamID64 together.

    python -m ladder.ufer_import "UFER.xlsx" --out data/seed/ufer.json
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:                                    # pragma: no cover
    sys.exit("openpyxl is missing:  python -m pip install openpyxl")


def _cell(row: tuple, idx: int):
    return row[idx] if idx < len(row) and row[idx] is not None else None


def read_ranking(wb) -> list[dict]:
    """Sheet `Ranking`: rank | delta | title | name | rating | +/-"""
    ws = wb["Ranking"]
    players: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _cell(row, 3)
        rating = _cell(row, 4)
        if not name or rating is None:
            continue
        if str(name).strip() in ("", "[NO_TEAMMATE]"):
            continue
        try:
            rating = float(rating)
        except (TypeError, ValueError):
            continue
        players.append({
            "name": str(name).strip(),
            "rating": round(rating, 1),
            "title": str(_cell(row, 2) or "").strip() or None,
            "rank": int(_cell(row, 0)) if isinstance(_cell(row, 0), (int, float)) else None,
            "steam_ids": [],          # filled in from recorded matches
        })
    return players


def read_peaks(wb) -> dict[str, float]:
    """Sheet `Players info`: name and peak rating.

    The sheet notes "peak not working yet", so the values are incomplete.
    They are imported anyway — a partial high is better than none — and from
    the import onwards `ratings.recompute` tracks the peak correctly itself.
    """
    if "Players info" not in wb.sheetnames:
        return {}
    ws = wb["Players info"]
    peaks: dict[str, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, peak = _cell(row, 3), _cell(row, 5)
        if not name or not isinstance(peak, (int, float)):
            continue
        peaks[str(name).strip()] = round(float(peak), 1)
    return peaks


def read_series(wb) -> list[dict]:
    """The rows present in the `Elo Algorithm` sheets, as events.

    There are few of them — those sheets are scratch space for the latest
    calculation — but they are the only place a series appears in full: both
    ratings, game count, score, date and event name. That is what the golden
    tests are built from.
    """
    events: list[dict] = []
    if "Elo Algorithm 1v1" in wb.sheetnames:
        for row in wb["Elo Algorithm 1v1"].iter_rows(min_row=2, values_only=True):
            a, b, ra, rb = (_cell(row, i) for i in (0, 1, 2, 3))
            games, score_a = _cell(row, 6), _cell(row, 7)
            if not a or not b or games is None or score_a is None:
                continue
            events.append({
                "kind": "1v1", "a": str(a).strip(), "b": str(b).strip(),
                "rating_a": ra, "rating_b": rb,
                "games": int(games), "score_a": int(score_a),
                "date": str(_cell(row, 11) or "")[:10],
                "event": str(_cell(row, 12) or "").strip() or None,
            })
    for sheet, coop in (("Elo Algorithm 2v2 TDM", False),
                        ("Elo Algorithm 3v3 TDM", False),
                        ("Elo Algorithm 2v2 Coop", True)):
        if sheet not in wb.sheetnames:
            continue
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            t1, t2 = _cell(row, 0), _cell(row, 1)
            games, score_a = _cell(row, 6), _cell(row, 7)
            if not t1 or not t2 or games is None or score_a is None:
                continue
            events.append({
                "kind": "team", "coop": coop,
                "team_a_label": str(t1).strip(), "team_b_label": str(t2).strip(),
                "rating_a": _cell(row, 2), "rating_b": _cell(row, 3),
                "games": int(games), "score_a": int(score_a),
                # Player names sit inside the +/- columns ("Name -28.2")
                # rather than in fields of their own, so mapping them to
                # teams stays manual.
                "raw_deltas": [str(_cell(row, i)) for i in range(9, 15)
                               if _cell(row, i) is not None],
                "sheet": sheet,
            })
    return events


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx", help="export of the ranking spreadsheet (.xlsx)")
    ap.add_argument("--out", default="data/seed/ufer.json")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    players = read_ranking(wb)
    peaks = read_peaks(wb)
    for p in players:
        if p["name"] in peaks:
            p["peak"] = peaks[p["name"]]
    series = read_series(wb)

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps({
        "source": Path(args.xlsx).name,
        "note": ("Seed from the ranking sheet. Display names without Steam "
                 "IDs; the link comes from recorded matches."),
        "players": players,
        "series": series,
    }, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"{len(players)} players, {len(series)} series -> {out}")
    if players:
        print("\nTop:")
        for p in players[:5]:
            print(f"   {p['rank']:>3}. {p['name']:<32} {p['rating']:>7.1f}  {p['title']}")
        print(f"\nwith a peak value: {sum('peak' in p for p in players)}/"
              f"{len(players)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
